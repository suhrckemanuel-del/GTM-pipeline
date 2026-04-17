"""
Export the outreach pipeline to an Excel workbook.

Output: pipeline/outreach_pipeline.xlsx
Sheets:
  1. Batch-01 Done  — companies whose first touch was 2026-04-11
  2. Batch-02 Next  — companies whose first touch is 2026-04-18
  3. Full Pipeline  — all companies in the pipeline

Columns per sheet:
  Company | Vertical | Stage | Headcount | Contact Name | Role | Group
  | LinkedIn URL | Email | Email Status | Send Date

Usage:
    python scripts/export_pipeline_xlsx.py
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install dependencies: pip install -r requirements.txt", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
PIPELINE = ROOT / "pipeline"

# Light alternating row fills for company groups
FILL_A = PatternFill("solid", fgColor="EFF3FB")  # soft blue-grey
FILL_B = PatternFill("solid", fgColor="FFFFFF")  # white
FILL_HEADER = PatternFill("solid", fgColor="2D5BE3")  # brand blue
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_COMPANY = Font(bold=True)

COLUMNS = [
    "Company",
    "Vertical",
    "Stage",
    "Headcount",
    "Contact Name",
    "Role",
    "Group",
    "LinkedIn URL",
    "Email",
    "Email Status",
    "Send Date",
]


def load_companies() -> dict[int, dict]:
    path = PIPELINE / "companies.csv"
    out: dict[int, dict] = {}
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cid = int(row["company_id"])
            out[cid] = row
    return out


def load_contacts() -> dict[int, dict]:
    """Returns contacts keyed by contact_id."""
    path = PIPELINE / "contacts.csv"
    out: dict[int, dict] = {}
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cid = int(row["contact_id"])
            out[cid] = row
    return out


def load_outreach() -> dict[int, dict]:
    """Returns first outreach touch keyed by contact_id."""
    path = PIPELINE / "outreach.csv"
    out: dict[int, dict] = {}
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cid = int(row["contact_id"])
            if cid not in out:  # keep only touch_number==1
                out[cid] = row
    return out


def batch_company_ids(outreach: dict[int, dict]) -> tuple[set[int], set[int]]:
    """Return (batch01_ids, batch02_ids) as sets of company_id ints."""
    b01: set[int] = set()
    b02: set[int] = set()
    for row in outreach.values():
        mv = row.get("message_version", "")
        cid = int(row["company_id"])
        if "batch-01" in mv:
            b01.add(cid)
        elif "batch-02" in mv:
            b02.add(cid)
    return b01, b02


def build_rows(
    company_ids: set[int] | None,
    companies: dict[int, dict],
    contacts: dict[int, dict],
    outreach: dict[int, dict],
) -> list[list]:
    """
    Build data rows for one sheet.
    company_ids=None means all companies.
    One row per contact, sorted by company_id then contact_id.
    """
    rows = []
    # Group contacts by company
    by_company: dict[int, list[tuple[int, dict]]] = {}
    for cid, contact in contacts.items():
        comp_id = int(contact["company_id"])
        if company_ids is not None and comp_id not in company_ids:
            continue
        by_company.setdefault(comp_id, []).append((cid, contact))

    for comp_id in sorted(by_company):
        comp = companies.get(comp_id, {})
        company_name = comp.get("name", "")
        vertical = comp.get("vertical", "")
        stage = comp.get("stage", "")
        headcount = comp.get("headcount_estimate", "")

        contact_list = sorted(by_company[comp_id], key=lambda x: x[0])
        for contact_id, contact in contact_list:
            ot = outreach.get(contact_id, {})
            send_date = ot.get("sent_date", "")
            rows.append([
                company_name,
                vertical,
                stage,
                headcount,
                contact.get("name", ""),
                contact.get("role", ""),
                contact.get("split_test_group", ""),
                contact.get("profile_url", ""),
                contact.get("work_email", ""),
                contact.get("email_status", ""),
                send_date,
            ])
    return rows


def write_sheet(
    ws,
    data_rows: list[list],
    sheet_title: str,
) -> None:
    # Header row
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Data rows with alternating colour per company group
    current_company = None
    fill_toggle = True
    for row_data in data_rows:
        company_name = row_data[0]
        if company_name != current_company:
            current_company = company_name
            fill_toggle = not fill_toggle
        fill = FILL_A if fill_toggle else FILL_B

        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        # Bold the company name on first contact row
        prev_row_company = ws.cell(row=row_idx - 1, column=1).value if row_idx > 2 else None
        if company_name != prev_row_company:
            ws.cell(row=row_idx, column=1).font = FONT_COMPANY

    # Auto-width columns
    col_widths = {i: len(COLUMNS[i - 1]) for i in range(1, len(COLUMNS) + 1)}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 0), min(len(str(cell.value)), 60)
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    ws.row_dimensions[1].height = 18
    ws.title = sheet_title


def main() -> None:
    companies = load_companies()
    contacts = load_contacts()
    outreach = load_outreach()
    b01_ids, b02_ids = batch_company_ids(outreach)

    wb = Workbook()

    # Sheet 1: Batch-01 Done
    ws1 = wb.active
    rows_b01 = build_rows(b01_ids, companies, contacts, outreach)
    write_sheet(ws1, rows_b01, "Batch-01 Done")

    # Sheet 2: Batch-02 Next
    ws2 = wb.create_sheet()
    rows_b02 = build_rows(b02_ids, companies, contacts, outreach)
    write_sheet(ws2, rows_b02, "Batch-02 Next")

    # Sheet 3: Full Pipeline
    ws3 = wb.create_sheet()
    rows_all = build_rows(None, companies, contacts, outreach)
    write_sheet(ws3, rows_all, "Full Pipeline")

    out_path = PIPELINE / "outreach_pipeline.xlsx"
    wb.save(out_path)

    print(f"Saved {out_path}")
    print(f"  Batch-01 Done : {len(rows_b01)} rows ({len(b01_ids)} companies)")
    print(f"  Batch-02 Next : {len(rows_b02)} rows ({len(b02_ids)} companies)")
    print(f"  Full Pipeline : {len(rows_all)} rows")


if __name__ == "__main__":
    main()
